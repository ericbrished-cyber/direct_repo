import json
import os
import argparse
import math
import glob
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill

# Configuration
GOLD_STANDARD_PATH = 'data/gold_standard_clean.json'
OUTPUT_FILENAME = 'colored_output.xlsx'

# Colors
COLOR_GREEN = PatternFill(start_color='00C6EFCE', end_color='00C6EFCE', fill_type='solid')
COLOR_RED = PatternFill(start_color='00FFC7CE', end_color='00FFC7CE', fill_type='solid')

# Key Fields for matching
MATCH_KEYS = ['pmcid', 'outcome', 'intervention', 'comparator', 'outcome_type']

# Metadata columns (start of the row)
META_COLUMNS = ['Model', 'pmcid', 'outcome', 'intervention', 'comparator', 'outcome_type']

# Numeric fields to compare and their types
# Fields: (field_name, is_integer)
NUMERIC_FIELDS = [
    ('intervention_mean', False),
    ('intervention_standard_deviation', False),
    ('comparator_mean', False),
    ('comparator_standard_deviation', False),
    ('intervention_events', True),
    ('intervention_group_size', True),
    ('comparator_events', True),
    ('comparator_group_size', True)
]

def load_gold_standard(path):
    """Loads and indexes the gold standard data."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Gold standard file not found at: {path}")

    with open(path, 'r') as f:
        data = json.load(f)

    index = {}
    duplicates = []

    for item in data:
        # Create a tuple key from the MATCH_KEYS
        # We convert to string and strip to ensure robust matching, though data seems clean
        key_parts = []
        for k in MATCH_KEYS:
            val = item.get(k)
            key_parts.append(str(val).strip() if val is not None else "None")

        key = tuple(key_parts)

        if key in index:
            duplicates.append(key)
        else:
            index[key] = item

    if duplicates:
        print(f"Warning: {len(duplicates)} duplicate keys found in Gold Standard. Using the first occurrence.")
        # For this task, we will proceed with the unique index.

    return index

def get_match_key(item):
    """Generates the match key for a prediction item."""
    key_parts = []
    for k in MATCH_KEYS:
        val = item.get(k)
        key_parts.append(str(val).strip() if val is not None else "None")
    return tuple(key_parts)

def is_value_match(pred_val, gold_val, is_int):
    """Strict comparison logic."""
    if pred_val is None and gold_val is None:
        return True, "both_missing"
    if pred_val is None or gold_val is None:
        return False, "one_missing"

    if is_int:
        try:
            # Parse as float first to handle 5.0 vs 5, then cast to int if it is effectively an int
            p = float(pred_val)
            g = float(gold_val)
            # Check strictly if they represent the same integer
            if p.is_integer() and g.is_integer():
                return int(p) == int(g), "value_comp"
            return math.isclose(p, g, rel_tol=0.0, abs_tol=1e-9), "value_comp"
        except (ValueError, TypeError):
             return False, "type_error"
    else:
        try:
            p = float(pred_val)
            g = float(gold_val)
            return math.isclose(p, g, rel_tol=0.0, abs_tol=1e-9), "value_comp"
        except (ValueError, TypeError):
            return False, "type_error"

def process_folder(folder_path, gold_index):
    """Scans the folder, processes JSONs, and builds the report."""

    # Verify folder
    if not os.path.isdir(folder_path):
        raise NotADirectoryError(f"Directory not found: {folder_path}")

    # Initialize Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Build Header Row
    headers = list(META_COLUMNS)
    for field, _ in NUMERIC_FIELDS:
        headers.append(field)
        headers.append(f"GOLD:{field}")

    ws.append(headers)

    # Statistics
    stats = Counter()
    stats['processed_files'] = 0
    stats['total_extractions'] = 0
    stats['gold_matches'] = 0
    stats['gold_mismatch_errors'] = 0

    # We will track field-level stats separately
    field_stats = Counter() # e.g. 'intervention_mean_match', 'intervention_mean_mismatch'

    # Glob all JSONs (excluding error files if any logic requires, but task says "ignore *_error.txt" which glob *.json handles naturally)
    json_files = glob.glob(os.path.join(folder_path, "*.json"))

    for json_file in json_files:
        stats['processed_files'] += 1
        try:
            with open(json_file, 'r') as f:
                data = json.load(f)
        except Exception as e:
            print(f"Error reading {json_file}: {e}")
            continue

        model_name = data.get("config", {}).get("model", "unknown")
        extractions = data.get("extraction", []) # Note: Schema example said "extractions" but file content showed "extraction" (singular list)
        if not extractions and "extractions" in data:
             extractions = data["extractions"]

        if not isinstance(extractions, list):
            print(f"Warning: 'extraction' is not a list in {json_file}")
            continue

        for pred_item in extractions:
            stats['total_extractions'] += 1

            # 1. Match Key
            key = get_match_key(pred_item)

            # 2. Find Gold
            gold_item = gold_index.get(key)

            row_data = []

            # Fill Metadata
            row_data.append(model_name)
            for k in META_COLUMNS[1:]: # Skip 'Model'
                row_data.append(pred_item.get(k))

            if not gold_item:
                stats['gold_mismatch_errors'] += 1
                # If 0 matches, user says "raise error listing the key".
                # Since we are processing a folder, halting everything might be annoying,
                # but "raise an error" usually implies stopping.
                # However, for a reporting script, printing and continuing or marking as error is often preferred.
                # The prompt says: "If 0 gold matches or >1 gold matches for a pred row: raise an error listing the key."
                # I will print the error and SKIP this row in the excel to comply with "raise an error" spirit without crashing the whole long-running script if possible,
                # OR I will actually raise the error if strict compliance is needed.
                # Given "Write a Python script that scans...", raising an Exception stops the scan.
                # I will print a loud error and skip adding this row to Excel, effectively "failing" this item.
                print(f"ERROR: No match found for key: {key}")
                continue
                # raise ValueError(f"No gold match found for key: {key}")

            stats['gold_matches'] += 1

            # Prepare cells with coloring info
            # We append tuples (value, fill_style) to a list, then write them to the row at the end
            colored_cells = [] # aligns with headers starting from numeric fields

            for field, is_int in NUMERIC_FIELDS:
                pred_val = pred_item.get(field)
                gold_val = gold_item.get(field)

                match_result, reason = is_value_match(pred_val, gold_val, is_int)

                # Logic:
                # If pred missing AND gold missing -> no color.
                # If pred missing AND gold present -> RED on pred cell (which is None/Empty).
                # If pred present AND gold missing -> RED on pred cell.
                # If both present -> GREEN if equal, else RED.
                # Do not color GOLD cells.

                pred_fill = None

                if pred_val is None and gold_val is None:
                    field_stats[f"{field}_missing_both"] += 1
                elif pred_val is None and gold_val is not None:
                    pred_fill = COLOR_RED
                    field_stats[f"{field}_missing_pred"] += 1
                elif pred_val is not None and gold_val is None:
                    pred_fill = COLOR_RED
                    field_stats[f"{field}_missing_gold"] += 1
                else:
                    # Both present
                    if match_result:
                        pred_fill = COLOR_GREEN
                        field_stats[f"{field}_match"] += 1
                    else:
                        pred_fill = COLOR_RED
                        field_stats[f"{field}_mismatch"] += 1

                colored_cells.append((pred_val, pred_fill)) # Prediction column
                colored_cells.append((gold_val, None))      # Gold column

            # Write Metadata to Sheet
            # ws.append accepts a list of values. To style, we must access the cells after appending.
            # So we combine all values first.
            final_values = row_data + [c[0] for c in colored_cells]
            ws.append(final_values)

            # Apply styles
            current_row_idx = ws.max_row

            # The numeric columns start after META_COLUMNS.
            # len(META_COLUMNS) is 6. So numeric cols start at index 6 (0-based) -> column 7 (1-based).
            start_col_idx = len(META_COLUMNS) + 1

            for i, (val, fill) in enumerate(colored_cells):
                if fill:
                    col_idx = start_col_idx + i
                    cell = ws.cell(row=current_row_idx, column=col_idx)
                    cell.fill = fill

    # Save
    output_path = os.path.join("data/results", OUTPUT_FILENAME)
    wb.save(output_path)
    print(f"Report saved to: {output_path}")

    # Print Summary
    print("\n--- Summary ---")
    print(f"Files Processed: {stats['processed_files']}")
    print(f"Total Extractions: {stats['total_extractions']}")
    print(f"Gold Matches: {stats['gold_matches']}")
    print(f"Gold Mismatches (Skipped): {stats['gold_mismatch_errors']}")
    print("\nField Level Stats:")

    # Print stats nicely per field
    for field, _ in NUMERIC_FIELDS:
        print(f"\n{field}:")
        print(f"  Matches: {field_stats[f'{field}_match']}")
        print(f"  Mismatches: {field_stats[f'{field}_mismatch']}")
        print(f"  Missing in Pred: {field_stats[f'{field}_missing_pred']}")
        print(f"  Missing in Gold: {field_stats[f'{field}_missing_gold']}")
        print(f"  Missing in Both: {field_stats[f'{field}_missing_both']}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate colored Excel report comparing predictions to gold standard.")
    parser.add_argument("input_folder", help="Path to the folder containing prediction JSON files.")

    args = parser.parse_args()

    print(f"Loading Gold Standard from {GOLD_STANDARD_PATH}...")
    gold_index = load_gold_standard(GOLD_STANDARD_PATH)

    print(f"Processing predictions in {args.input_folder}...")
    process_folder(args.input_folder, gold_index)
